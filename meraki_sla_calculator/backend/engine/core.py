########################################################
# .......Core module is made to build the logic....... #
# ...for calculation of the availability of devices... #
########################################################
from openpyxl import Workbook
from backend.engine.starters import *

presentTime = None


# Create a logic for calculating sla
def core_engine():
    """
    This method is the logic to calculate the availability of each device.
    :return: None
    """
    global presentTime
    presentTime = datetime.now(tz=pytz.timezone("UTC")).replace(tzinfo=None)
    try:
        logger.info("Execution started..")
        # Check if the SLA file exists in the /files directory
        # Update the org status file at every execution.

        # Load sla file in a variable to save r/w operation
        slaFile = read(SLA_JSON_FILE)

        # Create a loop to examine each device
        def get_device_sla(device):
            if device["serial"] in read(MAPPER_FILE):

                # device_timezone = pytz.timezone(net_dev_map_file[device["serial"]]["timeZone"])
                # Update reporting time of device at every execution.
                try:
                    lastReportTime = strp_dtime((device["lastReportedAt"]))
                    # If last reporting time is older than the engine's start time, overwrite it with the engine's start
                    # time
                    if lastReportTime < strp_dtime(slaFile[device["serial"]]["sla_start_time"]):
                        lastReportTime = slaFile[device["serial"]]["sla_start_time"]
                    else:
                        lastReportTime = str(strp_dtime((device["lastReportedAt"])))
                    slaFile[device["serial"]].update(last_report_time=lastReportTime)
                except (ConnectionError, AttributeError):
                    lastReportTime = slaFile[device["serial"]]["sla_start_time"]
                    slaFile[device["serial"]].update(last_report_time=lastReportTime)

                slaFile[device["serial"]].update(status=device["status"],
                                                 current_time=f"{presentTime}")

                # Maintenance Code
                for network in read(ORG_NETWORKS_FILE):
                    if network["id"] == slaFile[device["serial"]]["networkId"]:

                        if "Maintenance" in network["tags"]:
                            slaFile[device["serial"]].update(last_report_time=str(presentTime))
                            slaFile[device["serial"]].update(in_maintenance="true")
                            slaFile[device["serial"]].update(last_report_time_mti=str(presentTime))
                        else:
                            slaFile[device["serial"]].update(in_maintenance="false")

                try:
                    mti = slaFile[device["serial"]]["last_report_time_mti"]
                    lrt = slaFile[device["serial"]]["last_report_time"]
                    if strp_dtime(mti) >= strp_dtime(lrt):
                        slaFile[device["serial"]].update(last_report_time=mti)
                    else:
                        slaFile[device["serial"]].update(last_report_time=lrt)
                except KeyError:
                    pass
                # Maintenance code end

                # Calculate the downtime of each device if they are offline.
                if slaFile[device["serial"]]["last_report_time"] == slaFile[device["serial"]]["reporting_time_tracker"]:

                    if device["status"] == "online":
                        logger.warning(f"{device['name']} reporting failed but status is ONLINE on dashboard. "
                                       f"Forcing it to be OFFLINE.")
                        previous_log = device
                        previous_log["lastReportedAt"] = slaFile[device["serial"]]["last_report_time"]
                        logger.info("Before")
                        logger.info(previous_log)
                        logger.info("After")
                        logger.info(device)
                        slaFile[device["serial"]].update(status="offline(forced)")

                    elif device["status"] == "offline":
                        pass
                    lastDownSeconds = slaFile[device["serial"]]["last_downtime_seconds"]
                    lastReportTime = strp_dtime(slaFile[device["serial"]]["last_report_time"])
                    # Formula: Previous downTime + Present downTime
                    downTime = lastDownSeconds + (presentTime - lastReportTime).total_seconds()
                    slaFile[device["serial"]].update(total_downtime_seconds=downTime)

                elif slaFile[device["serial"]]["last_report_time"] != \
                        slaFile[device["serial"]]["reporting_time_tracker"]:
                    if device["status"] == "offline" or device["status"] == "alerting":
                        logger.warning(f"{device['name']} reporting successful but status is OFFLINE on dashboard. "
                                       f"Forcing it to be ONLINE.")
                        slaFile[device["serial"]].update(status="online(forced)")

                    elif device["status"] == "online":
                        # If device is online copy the latest downTime in another attribute "last_downtime_seconds"
                        # for future calculations
                        pass
                    lastTotalDownSeconds = slaFile[device["serial"]]["total_downtime_seconds"]
                    latestReportingTime = slaFile[device["serial"]]["last_report_time"]
                    slaFile[device["serial"]].update(last_downtime_seconds=lastTotalDownSeconds)
                    slaFile[device["serial"]].update(reporting_time_tracker=latestReportingTime)

                # Calculate the sla of each device using percentage formula.
                # To calculate we need 2 time intervals
                # a) Duration between the device added till present time.
                # b) DownTime duration
                deviceSlaStartTimer = strp_dtime(slaFile[device["serial"]]["sla_start_time"])
                totalSeconds = (presentTime - deviceSlaStartTimer).total_seconds()
                totalDownSeconds = slaFile[device["serial"]].get("total_downtime_seconds")
                slaFile[device["serial"]].update(total_time_in_seconds=totalSeconds)
                # Apply percentage formula
                try:
                    if slaFile[device["serial"]]["in_maintenance"] == "true":
                        previous_sla = slaFile[device["serial"]]["sla"]
                        slaFile[device["serial"]].update(sla=previous_sla)

                    else:
                        slaUpTime = round(((totalSeconds - totalDownSeconds) / totalSeconds) * 100, 2)
                        slaFile[device["serial"]].update(sla=slaUpTime)
                except ZeroDivisionError:
                    slaFile[device["serial"]].update(sla=0)

                try:
                    if slaFile[device["serial"]]["in_maintenance"] == "true":
                        slaFile[device["serial"]].update(status="online(in maintenance)")
                except KeyError:
                    pass

        write(ORG_DEVICE_STATUS_FILE, get_org_devices_status_of_all_dashboard())

        with concurrent.futures.ThreadPoolExecutor() as executor:
            logger.info("running the logic..")
            thread = [executor.submit(get_device_sla, device) for device in read(ORG_DEVICE_STATUS_FILE)]
            for output in concurrent.futures.as_completed(thread):
                pass
        """for device in read(ORG_DEVICE_STATUS_FILE):
            get_device_sla(device)"""
        write(SLA_JSON_FILE, slaFile)
        logger.info("Execution completed.")
        generate_single_report()
    except (ConnectionError,) as e:
        logger.exception("Error occurred: skipping this execution cycle.")


# Create a logic for calculating sla
def core_engine_for_DB():
    """
    This method is the logic to calculate the availability of each device.
    :return: None
    """
    global presentTime
    presentTime = datetime.now(tz=pytz.timezone("UTC")).replace(tzinfo=None)
    try:
        logger.info("Execution started..")

        # Check if the SLA file exists in the /files directory
        # Update the org status file at every execution.

        # Load sla file in a variable to save r/w operation

        # Create a loop to examine each device
        def get_device_sla(device):
            if device_in_mapper_already_exists(device["serial"]):
                if SLA_COLLECTION.find_one({"serial": device["serial"]}) is None:
                    check_new_device_for_DB()

                device_in_sla_db: dict = SLA_COLLECTION.find_one({"serial": device["serial"]})

                # device_timezone = pytz.timezone(net_dev_map_file[device["serial"]]["timeZone"])
                # Update reporting time of device at every execution.
                try:
                    lastReportTime = strp_dtime((device["lastReportedAt"]))
                    # If last reporting time is older than the engine's start time, overwrite it with the engine's start
                    # time
                    if lastReportTime < strp_dtime(device_in_sla_db["sla_start_time"]):
                        lastReportTime = device_in_sla_db["sla_start_time"]
                    else:
                        lastReportTime = str(strp_dtime((device["lastReportedAt"])))
                    device_in_sla_db.update(last_report_time=lastReportTime)
                except (ConnectionError, AttributeError):
                    lastReportTime = device_in_sla_db["sla_start_time"]
                    device_in_sla_db.update(last_report_time=lastReportTime)

                device_in_sla_db.update(status=device["status"],
                                        current_time=f"{presentTime}")

                # Maintenance code for DB
                for network in NETWORKS_COLLECTION.find():
                    if network["networkId"] == device_in_sla_db["networkId"]:
                        if "Maintenance" in network["tags"]:
                            device_in_sla_db.update(last_report_time=str(presentTime))
                            device_in_sla_db.update(in_maintenance="true")
                            device_in_sla_db.update(last_report_time_mti=str(presentTime))
                        else:
                            device_in_sla_db.update(in_maintenance="false")

                try:
                    mti = device_in_sla_db["last_report_time_mti"]
                    lrt = device_in_sla_db["last_report_time"]
                    if strp_dtime(mti) >= strp_dtime(lrt):
                        device_in_sla_db.update(last_report_time=mti)
                    else:
                        device_in_sla_db.update(last_report_time=lrt)
                except KeyError:
                    pass
                # Maintenance code end.

                # Calculate the downtime of each device if they are offline.
                if device_in_sla_db["last_report_time"] == device_in_sla_db["reporting_time_tracker"]:

                    if device["status"] == "online":
                        logger.warning(f"{device['name']} reporting failed but status is ONLINE on dashboard. "
                                       f"Forcing it to be OFFLINE.")
                        device_in_sla_db.update(status="offline(forced)")

                    elif device["status"] == "offline":
                        pass
                    lastDownSeconds = device_in_sla_db["last_downtime_seconds"]
                    lastReportTime = strp_dtime(device_in_sla_db["last_report_time"])
                    # Formula: Previous downTime + Present downTime
                    downTime = lastDownSeconds + (presentTime - lastReportTime).total_seconds()
                    device_in_sla_db.update(total_downtime_seconds=downTime)

                elif device_in_sla_db["last_report_time"] != \
                        device_in_sla_db["reporting_time_tracker"]:
                    if device["status"] == "offline" or device["status"] == "alerting":
                        logger.warning(f"{device['name']} reporting successful but status is OFFLINE on dashboard. "
                                       f"Forcing it to be ONLINE.")
                        device_in_sla_db.update(status="online(forced)")

                    elif device["status"] == "online":
                        # If device is online copy the latest downTime in another attribute "last_downtime_seconds"
                        # for future calculations
                        pass
                    lastTotalDownSeconds = device_in_sla_db["total_downtime_seconds"]
                    latestReportingTime = device_in_sla_db["last_report_time"]
                    device_in_sla_db.update(last_downtime_seconds=lastTotalDownSeconds)
                    device_in_sla_db.update(reporting_time_tracker=latestReportingTime)

                # Calculate the sla of each device using percentage formula.
                # To calculate we need 2 time intervals
                # a) Duration between the device added till present time.
                # b) DownTime duration
                deviceSlaStartTimer = strp_dtime(device_in_sla_db["sla_start_time"])
                totalSeconds = (presentTime - deviceSlaStartTimer).total_seconds()
                totalDownSeconds = device_in_sla_db["total_downtime_seconds"]
                device_in_sla_db.update(total_time_in_seconds=totalSeconds)
                # Apply percentage formula
                try:
                    if device_in_sla_db["in_maintenance"] == "true":
                        previous_sla = device_in_sla_db["sla"]
                        device_in_sla_db.update(sla=previous_sla)
                    else:
                        slaUpTime = round(((totalSeconds - totalDownSeconds) / totalSeconds) * 100, 2)
                        device_in_sla_db.update(sla=slaUpTime)
                except ZeroDivisionError:
                    device_in_sla_db.update(sla=0)

                try:
                    if device_in_sla_db["in_maintenance"] == "true":
                        device_in_sla_db.update(status="online(in maintenance)")
                except KeyError:
                    pass

                SLA_COLLECTION.find_one_and_update({"serial": device_in_sla_db["serial"]}, {"$set": device_in_sla_db})
        org_devices_status = get_org_devices_status_of_all_dashboard_for_DB()
        """with concurrent.futures.ThreadPoolExecutor() as executor:
            logger.info("running the logic..")
            thread = [executor.submit(get_device_sla, device) for device in org_devices_status]
            for output in concurrent.futures.as_completed(thread):
                pass"""
        for device in org_devices_status:
            get_device_sla(device)
        logger.info("Executing completed.")
        generate_single_report_for_DB()

    except (ConnectionError,) as e:
        logger.exception("Error occurred: skipping this execution cycle.")


def generate_single_report():
    """
    This function will create a report after the availability calculation is completed.
    :return:
    """
    try:
        logger.info("Updating the single availability report..")
        excelReport = Workbook()
        slaFile = read(SLA_JSON_FILE)
        ROW = 2
        # The sheet 1 of the workbook will contain the time when the script started
        default_sheet = excelReport.get_sheet_by_name('Sheet')
        sheet = excelReport.create_sheet(title="Availability")

        create_header_cell(row=1, column=1, value="Serial No.", target_sheet=sheet)
        create_header_cell(row=1, column=2, value="Name", target_sheet=sheet)
        create_header_cell(row=1, column=3, value="Mac address", target_sheet=sheet)
        create_header_cell(row=1, column=4, value="Model", target_sheet=sheet)
        create_header_cell(row=1, column=5, value="Network", target_sheet=sheet)
        create_header_cell(row=1, column=6, value="Sla(%)", target_sheet=sheet)
        create_header_cell(row=1, column=7, value="Downtime(mins)", target_sheet=sheet)
        create_header_cell(row=1, column=8, value="Total Time(in mins", target_sheet=sheet)
        create_header_cell(row=1, column=9, value="Start Time(IST)", target_sheet=sheet)
        create_header_cell(row=1, column=10, value="End Time(IST)", target_sheet=sheet)

        for device, attribute in slaFile.items():
            if device == "script_start_time":
                pass
            create_data_cell(row=ROW, column=1, value=device, target_sheet=sheet)
            create_data_cell(row=ROW, column=2, value=attribute["name"], target_sheet=sheet)
            create_data_cell(row=ROW, column=3, value=attribute["mac"], target_sheet=sheet)
            create_data_cell(row=ROW, column=4, value=attribute["model"], target_sheet=sheet)
            create_data_cell(row=ROW, column=5, value=attribute["network"], target_sheet=sheet)
            create_data_cell(row=ROW, column=6, value=attribute["sla"], target_sheet=sheet)
            create_data_cell(row=ROW, column=7, value=round(attribute['total_downtime_seconds'] / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=8, value=round(attribute['total_time_in_seconds'] / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=9, value=strp_dtime_UTC_to_IST(attribute['sla_start_time']),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=10, value=strp_dtime_UTC_to_IST(attribute['current_time']),
                             target_sheet=sheet)
            ROW += 1
        excelReport.remove(excelReport['Sheet'])

        # Network Wise SLA Calculation
        network_sla_sheet = excelReport.create_sheet(title="Network-Wise")
        ROW = 1
        LAN_AVG = []
        WIFI_AVG = []
        create_header_cell(row=ROW, column=1, value="Network", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=2, value="Lan Availability(%)", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=3, value="Wi-Fi Availability(%)", target_sheet=network_sla_sheet)

        for network in read(ORG_NETWORKS_FILE):
            if "0-Active" in network["tags"]:
                ROW += 1
                create_data_cell(row=ROW, column=1, value=network["name"], target_sheet=network_sla_sheet)
                for device, attributes in slaFile.items():
                    if attributes["networkId"] == network["id"]:
                        if attributes["model"].find("MS") != -1:
                            LAN_AVG.append(attributes["sla"])
                        elif attributes["model"].find("MR") != -1:
                            WIFI_AVG.append(attributes["sla"])

                LAN_ADDITION = 0
                WIFI_ADDITION = 0
                for sla_entry in LAN_AVG:
                    LAN_ADDITION = LAN_ADDITION + sla_entry

                for sla_entry in WIFI_AVG:
                    WIFI_ADDITION = WIFI_ADDITION + sla_entry

                try:
                    LAN_SLA = round(LAN_ADDITION / len(LAN_AVG), 2)
                    WIFI_SLA = round(WIFI_ADDITION / len(WIFI_AVG), 2)
                except ZeroDivisionError:
                    LAN_SLA = 0
                    WIFI_SLA = 0

                create_data_cell(row=ROW, column=2, value=LAN_SLA, target_sheet=network_sla_sheet)
                create_data_cell(row=ROW, column=3, value=WIFI_SLA, target_sheet=network_sla_sheet)
                LAN_AVG = []
                WIFI_AVG = []
        try:
            excelReport.save(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}")
            logger.info("Availability Report updated successfully..")
        except PermissionError:
            logger.warning("PermissionError: Report is open. It will now be updated in the next execution")
    except ConnectionError:
        logger.exception("Error while creating single availability report.")


def generate_single_report_for_DB():
    """
    This function will create a report after the availability calculation is completed.
    :return:
    """
    try:
        logger.info("Updating the single availability report..")
        excelReport = Workbook()
        # slaFile = read(SLA_JSON_FILE)
        ROW = 2
        # The sheet 1 of the workbook will contain the time when the script started
        default_sheet = excelReport.get_sheet_by_name('Sheet')
        sheet = excelReport.create_sheet(title="Availability")

        create_header_cell(row=1, column=1, value="Serial No.", target_sheet=sheet)
        create_header_cell(row=1, column=2, value="Name", target_sheet=sheet)
        create_header_cell(row=1, column=3, value="Mac address", target_sheet=sheet)
        create_header_cell(row=1, column=4, value="Model", target_sheet=sheet)
        create_header_cell(row=1, column=5, value="Network", target_sheet=sheet)
        create_header_cell(row=1, column=6, value="Sla(%)", target_sheet=sheet)
        create_header_cell(row=1, column=7, value="Downtime(mins)", target_sheet=sheet)
        create_header_cell(row=1, column=8, value="Total Time(in mins", target_sheet=sheet)
        create_header_cell(row=1, column=9, value="Start Time(IST)", target_sheet=sheet)
        create_header_cell(row=1, column=10, value="End Time(IST)", target_sheet=sheet)

        for device in SLA_COLLECTION.find():
            if device == "script_start_time":
                pass
            create_data_cell(row=ROW, column=1, value=device["serial"], target_sheet=sheet)
            create_data_cell(row=ROW, column=2, value=device["name"], target_sheet=sheet)
            create_data_cell(row=ROW, column=3, value=device["mac"], target_sheet=sheet)
            create_data_cell(row=ROW, column=4, value=device["model"], target_sheet=sheet)
            create_data_cell(row=ROW, column=5, value=device["network"], target_sheet=sheet)
            create_data_cell(row=ROW, column=6, value=device["sla"], target_sheet=sheet)
            create_data_cell(row=ROW, column=7, value=round(device['total_downtime_seconds'] / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=8, value=round(device['total_time_in_seconds'] / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=9, value=strp_dtime_UTC_to_IST(device['sla_start_time']),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=10, value=strp_dtime_UTC_to_IST(device['current_time']),
                             target_sheet=sheet)
            ROW += 1
        excelReport.remove(excelReport['Sheet'])

        # Network Wise SLA Calculation
        network_sla_sheet = excelReport.create_sheet(title="Network-Wise")
        ROW = 1
        LAN_AVG = []
        WIFI_AVG = []
        create_header_cell(row=ROW, column=1, value="Network", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=2, value="Lan Availability(%)", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=3, value="Wi-Fi Availability(%)", target_sheet=network_sla_sheet)

        for network in NETWORKS_COLLECTION.find():
            # if network["tags"].find("0-Active") != -1:
            ROW += 1
            create_data_cell(row=ROW, column=1, value=network["name"], target_sheet=network_sla_sheet)
            for device in SLA_COLLECTION.find():
                if device["networkId"] == network["networkId"]:
                    if device["model"].find("MS") != -1:
                        LAN_AVG.append(device["sla"])
                    elif device["model"].find("MR") != -1:
                        WIFI_AVG.append(device["sla"])

            LAN_ADDITION = 0
            WIFI_ADDITION = 0
            for sla_entry in LAN_AVG:
                LAN_ADDITION = LAN_ADDITION + sla_entry

            for sla_entry in WIFI_AVG:
                WIFI_ADDITION = WIFI_ADDITION + sla_entry

            try:
                LAN_SLA = round(LAN_ADDITION / len(LAN_AVG), 2)
                WIFI_SLA = round(WIFI_ADDITION / len(WIFI_AVG), 2)
            except ZeroDivisionError:
                LAN_SLA = 0
                WIFI_SLA = 0

            create_data_cell(row=ROW, column=2, value=LAN_SLA, target_sheet=network_sla_sheet)
            create_data_cell(row=ROW, column=3, value=WIFI_SLA, target_sheet=network_sla_sheet)
            LAN_AVG = []
            WIFI_AVG = []
        try:
            excelReport.save(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}")
            logger.info("Availability Report updated successfully..")
        except PermissionError:
            logger.warning("PermissionError: Report is open. It will now be updated in the next execution")
    except ConnectionError:
        logger.exception("Error while creating single availability report.")
