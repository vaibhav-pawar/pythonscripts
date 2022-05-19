###################################################
# ...........Maintenance Window API.............. #
###################################################
from backend.engine.helpers import *
from openpyxl.workbook import Workbook


def custom_response(status, code, msg, **kwargs):
    response = {"status": status, "code": code, "msg": msg}
    for k, v in kwargs.items():
        response.update({k: v})
    return response


def create_maintenance_window():
    # get networkId
    networkId = cinput("Enter network id: ")

    # validate the maintenance window template
    validation, window = validate_window(networkId, prepare_window_template())
    if not validation:
        return custom_response(status="error", code=-1, msg="Maintenance window already exist for given start time",
                               window=window)

    # if validated create an empty list and add new window
    maintenance_windows_list = [window]

    # if there are previous windows then append new window to the list of previous windows.
    if get_maintenance_windows(networkId) is not None:
        # get list of previous windows
        maintenance_windows: list = get_maintenance_windows(networkId)

        # add both lists:
        maintenance_windows_list = maintenance_windows_list + maintenance_windows
        # Update the db.
        NETWORK_MAINTENANCE_COLLECTION.find_one_and_update({"networkId": networkId},
                                                           {"$set": {"maintenance_windows": maintenance_windows_list}})

    # if there are NO previous windows then add new window to the list.
    else:
        NETWORK_MAINTENANCE_COLLECTION.find_one_and_update({"networkId": networkId},
                                                           {"$set": {"maintenance_windows": maintenance_windows_list}})

    return custom_response(status="success", code=0, msg="Maintenance window created successfully")


def validate_window(networkId, new_window_template):
    # get maintenance windows for given nid.
    windows = get_maintenance_windows(networkId)

    # compare new window with older windows. if exist the return False else return True.
    for window in windows:
        window_start_time = strp_dtime(f"{window['from_date']} {window['from_time']}")
        window_end_time = strp_dtime(f"{window['to_date']} {window['to_time']}")
        new_window_start_time = strp_dtime(f"{new_window_template['from_date']} {new_window_template['from_time']}")
        new_window_end_time = strp_dtime(f"{new_window_template['to_date']} {new_window_template['to_time']}")

        if window_start_time.timestamp() < new_window_start_time.timestamp() < window_end_time.timestamp():
            print("A Maintenance window is already created in this time frame.")
            print("Window: ", window)
            return False, window
        else:
            return True, new_window_template


def remove_maintenance_window(networkId, mid):
    windows: list = get_maintenance_windows(networkId)
    for window in windows:
        if window["mid"] == mid:
            windows.remove(window)
    NETWORK_MAINTENANCE_COLLECTION.find_one_and_update({"networkId": networkId},
                                                       {"$set": {"maintenance_windows": windows}})
    return get_maintenance_windows(networkId)


def get_maintenance_windows(networkId):
    network: dict = NETWORK_MAINTENANCE_COLLECTION.find_one({"networkId": networkId})
    try:
        return network["maintenance_windows"]
    except KeyError:
        return None


def get_maintenance_window(networkId, mid):
    maintenance_windows = get_maintenance_windows(networkId)
    try:
        for window in maintenance_windows:
            if window["mid"] == mid:
                return window
    except KeyError:
        return None


def prepare_window_template():
    mid = get_unique_id()
    from_date = cinput("Enter from date [yyyy-mm-dd]: ")
    to_date = cinput("Enter to date [yyyy-mm-dd]: ")
    from_time = cinput("Enter from time [hh:mm:ss]: ")
    to_time = cinput("Enter to time [hh:mm:ss]: ")
    comment = cinput("Enter Comment: ")
    window_start_time = strp_dtime(f"{from_date} {from_time}")
    window_end_time = strp_dtime(f"{to_date} {to_time}")

    if window_start_time.timestamp() > window_end_time.timestamp():
        return custom_response(status="error", code=-1, msg="Start and End DateTime incorrect.")
    maintenance_template = {
        "mid": mid,
        "from_date": from_date,
        "from_time": from_time,
        "to_date": to_date,
        "to_time": to_time,
        "comment": comment
    }
    return maintenance_template


def revise_total_downtime(window, downtime, sla_start_time: datetime, current_time: datetime):
    print("Previous DownTime: ", downtime)
    present_timestamp = current_time.timestamp()
    present_time = datetime.fromtimestamp(present_timestamp)
    window_start_time = strp_dtime(f"{window['from_date']} {window['from_time']}")
    window_end_time = strp_dtime(f"{window['to_date']} {window['to_time']}")
    print("Present Time: ", present_time)
    print("Window start time: ", window_start_time)

    # check if window start time is greater than start time if not window is of previous months
    if window_start_time.timestamp() > sla_start_time.timestamp():
        # print("window start time is greater than start time")
        # if window start time is less than present time if not then it is for future
        if window_start_time.timestamp() < present_timestamp:
            # print("window start time is less than present time")
            # if window end time is less then present time then maintenance window is over
            if window_end_time.timestamp() < present_timestamp:
                # print("window end time is less then present time")
                revise_downtime = downtime - int(window_end_time.timestamp() - window_start_time.timestamp())
                # after subtracting the maintenance window if revise time is less than zero than make it zero
                # print("Revise Time: ", revise_downtime)
                if revise_downtime < 0:
                    # print("Revise time is less then 0. Returning 0")
                    return 0
                else:
                    return revise_downtime
            # if window end time is greater than present time than maintenance window is ongoing
            elif window_end_time.timestamp() > present_timestamp:
                # print("window end time is greater than present time. Maintenance window is ongoing")
                revise_downtime = downtime - int(window_start_time.timestamp() + present_timestamp)
                # after subtracting the maintenance window if revise time is less than zero than make it zero
                # print("Revise Time: ", revise_downtime)
                if revise_downtime < 0:
                    # print("Revise time is less then 0. Returning 0")
                    return 0
                else:
                    return revise_downtime
            else:
                return downtime
        else:
            # print("window start time is greater than present time")
            return downtime
    else:
        # print("window start time is less than start time")
        return 0


def revise_total_downtime_addition(networkId, downtime, sla_start_time: datetime, current_time, device_name):
    print("------------------------------------------------------------------------")
    print("Network ID: ", networkId)
    print("Device name: ", device_name)
    windows = get_maintenance_windows(networkId)
    revise_time = 0
    if windows is not None:
        for window in windows:
            print("Window: ", window)
            print("Revise Time: ", revise_time)
            new_time = revise_total_downtime(window, downtime, sla_start_time, current_time)
            print("New Revise time from function: ", new_time, f" {revise_time} + {new_time}")
            revise_time += new_time
            print("Revise Time after adding: ", revise_time)
            print()
        print("------------------------------------------------------------------------")

        return revise_time
    else:
        print("------------------------------------------------------------------------")
        return downtime


def revised_sla(total_seconds, downtime):
    try:
        slaUpTime = round(((total_seconds - downtime) / total_seconds) * 100, 2)
        return slaUpTime
    except ZeroDivisionError:
        return 0


def generate_report_with_maintenance_filter():
    """
    This function will create a report after the availability calculation is completed.
    :return:
    """
    try:
        logger.info("Updating the single availability report..")
        excelReport = Workbook()
        # slaFile = read(SLA_JSON_FILE)
        ROW = 2
        # The sheet 1 of the workbook will contain the time when the script started
        default_sheet = excelReport.get_sheet_by_name('Sheet')
        sheet = excelReport.create_sheet(title="Availability")

        create_header_cell(row=1, column=1, value="Serial No.", target_sheet=sheet)
        create_header_cell(row=1, column=2, value="Name", target_sheet=sheet)
        create_header_cell(row=1, column=3, value="Mac address", target_sheet=sheet)
        create_header_cell(row=1, column=4, value="Model", target_sheet=sheet)
        create_header_cell(row=1, column=5, value="Network", target_sheet=sheet)
        create_header_cell(row=1, column=6, value="Sla(%)", target_sheet=sheet)
        create_header_cell(row=1, column=7, value="Downtime(mins)", target_sheet=sheet)
        create_header_cell(row=1, column=8, value="Total Time(in mins)", target_sheet=sheet)
        create_header_cell(row=1, column=9, value="Start Time(IST)", target_sheet=sheet)
        create_header_cell(row=1, column=10, value="End Time(IST)", target_sheet=sheet)

        for device in SLA_COLLECTION.find():
            if device == "script_start_time":
                pass
            revise_downtime = revise_total_downtime_addition(networkId=device["networkId"],
                                                             downtime=device["total_downtime_seconds"],
                                                             sla_start_time=strp_dtime_UTC_to_IST(
                                                                 device["sla_start_time"]),
                                                             current_time=strp_dtime_UTC_to_IST(device["current_time"]),
                                                             device_name=device["name"])
            create_data_cell(row=ROW, column=1, value=device["serial"], target_sheet=sheet)
            create_data_cell(row=ROW, column=2, value=device["name"], target_sheet=sheet)
            create_data_cell(row=ROW, column=3, value=device["mac"], target_sheet=sheet)
            create_data_cell(row=ROW, column=4, value=device["model"], target_sheet=sheet)
            create_data_cell(row=ROW, column=5, value=device["network"], target_sheet=sheet)

            revise_sla = revised_sla(total_seconds=device["total_time_in_seconds"],
                                     downtime=revise_downtime)
            create_data_cell(row=ROW, column=6, value=revise_sla, target_sheet=sheet)

            create_data_cell(row=ROW, column=7, value=round(revise_downtime / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=8, value=round(device['total_time_in_seconds'] / 60, 2),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=9, value=strp_dtime_UTC_to_IST(device['sla_start_time']),
                             target_sheet=sheet)
            create_data_cell(row=ROW, column=10, value=strp_dtime_UTC_to_IST(device['current_time']),
                             target_sheet=sheet)
            ROW += 1
        excelReport.remove(excelReport['Sheet'])

        # Network Wise SLA Calculation
        network_sla_sheet = excelReport.create_sheet(title="Network-Wise")
        ROW = 1
        LAN_AVG = []
        WIFI_AVG = []
        create_header_cell(row=ROW, column=1, value="Network", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=2, value="Lan Availability(%)", target_sheet=network_sla_sheet)
        create_header_cell(row=ROW, column=3, value="Wi-Fi Availability(%)", target_sheet=network_sla_sheet)

        for network in NETWORKS_COLLECTION.find():
            # if network["tags"].find("0-Active") != -1:
            ROW += 1
            create_data_cell(row=ROW, column=1, value=network["name"], target_sheet=network_sla_sheet)
            for device in SLA_COLLECTION.find():
                if device["networkId"] == network["networkId"]:
                    print(device["networkId"])
                    revise_downtime = \
                        revise_total_downtime_addition(networkId=device["networkId"],
                                                       downtime=device["total_downtime_seconds"],
                                                       sla_start_time=strp_dtime_UTC_to_IST(device["sla_start_time"]),
                                                       current_time=strp_dtime_UTC_to_IST(device["current_time"]),
                                                       device_name=device["name"])
                    print(revise_downtime)
                    revise_sla = revised_sla(total_seconds=device["total_time_in_seconds"],
                                             downtime=revise_downtime)
                    print(revise_sla)

                    if device["model"].find("MS") != -1:
                        LAN_AVG.append(revise_sla)
                    elif device["model"].find("MR") != -1:
                        WIFI_AVG.append(revise_sla)

            LAN_ADDITION = 0
            WIFI_ADDITION = 0
            for sla_entry in LAN_AVG:
                LAN_ADDITION = LAN_ADDITION + sla_entry

            for sla_entry in WIFI_AVG:
                WIFI_ADDITION = WIFI_ADDITION + sla_entry

            try:
                LAN_SLA = round(LAN_ADDITION / len(LAN_AVG), 2)
                WIFI_SLA = round(WIFI_ADDITION / len(WIFI_AVG), 2)
            except ZeroDivisionError:
                LAN_SLA = 0
                WIFI_SLA = 0

            create_data_cell(row=ROW, column=2, value=LAN_SLA, target_sheet=network_sla_sheet)
            create_data_cell(row=ROW, column=3, value=WIFI_SLA, target_sheet=network_sla_sheet)
            LAN_AVG = []
            WIFI_AVG = []
        try:
            excelReport.save(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_WITH_MAINTENANCE_NAME}")
            logger.info("Availability Report updated successfully..")
        except PermissionError:
            logger.warning("PermissionError: Report is open. It will now be updated in the next execution")
    except ConnectionError:
        logger.exception("Error while creating single availability report.")


nid = "L_646829496481108061"

print(jprint(get_maintenance_windows(nid)))
generate_report_with_maintenance_filter()
