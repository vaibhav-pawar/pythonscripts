########################################################
# ...........Module for Helper functions.............. #
# ..that will be used in rest of the code frequently.. #
########################################################
import concurrent.futures
import json
import shutil
import uuid
from datetime import datetime
import requests
from openpyxl.styles import Font, PatternFill, Side, Border
from pytz import timezone
from backend.baseConfig.config import *
from backend.engine.gmail_notifications import send_notification_email
from cryptography.fernet import Fernet
from backend.baseConfig.html_templates import SEND_REPORT_CONTENT


# Check if the required directories exists in the project folder. If not create it.
def create_paths():
    if not os.path.exists(LOGS_PATH):
        os.mkdir(LOGS_PATH)

    if not os.path.exists(FILES_PATH):
        os.mkdir(FILES_PATH)

    if not os.path.exists(REPORT_PATH):
        os.mkdir(REPORT_PATH)

    if not os.path.exists(MONTHLY_REPORT_PATH):
        os.mkdir(MONTHLY_REPORT_PATH)

    if not os.path.exists(DAILY_REPORT_PATH):
        os.mkdir(DAILY_REPORT_PATH)

    if not os.path.exists(LIVE_REPORT_PATH):
        os.mkdir(LIVE_REPORT_PATH)


def jprint(content):
    """
    This method will print the output in Pretty JSON format.
    :param content: define content/data/output that needs to be formatted.
    :return: JSON formatted output.
    """
    return json.dumps(content, indent=4)


def write(filename, content, mode="w", _json=True):
    """
    This method will write all contents in a file.
    :param filename: Specify the name of the file.
    :param content: Specify the content to write in a file.
    :param mode: Specify the mode. "w" means write "r" means read and "a" means append.
    :param _json: If bool JSON is true then it will write in a Pretty JSON format. If False then plain text strings.
    :return: None
    """
    try:
        with open(filename, mode) as f:
            if _json:
                f.write(json.dumps(content, indent=4))
            else:
                f.write(f"{content}")
    except PermissionError:
        logger.error("PermissionError, File is open")


def read(filename, _json=True):
    """
    This method will read all contents in a file.
    :param filename: Specify the name of the file.
    :param _json: If bool JSON is true it will read as JSON content else it will read as plain text content.
    :return: File contents.
    """
    with open(filename, "r") as f:
        content = f.read()
        if _json:
            return json.loads(content)
        else:
            return content


def strp_dtime(str_time):
    """
    This method will convert a string.datetime to datetime.datetime
    :param str_time: Specify the datetime in a string format.
    :return: Date and time in  a datetime format.
    """
    str_time = str_time.strip("Z").replace("T", " ")
    try:
        dtime = datetime.strptime(str_time, "%Y-%m-%d %H:%M:%S.%f")
    except ValueError:
        dtime = datetime.strptime(str_time, "%Y-%m-%d %H:%M:%S")

    return dtime


def strp_dtime_UTC_to_IST(str_time):
    """
    This method will convert a string.datetime to datetime.datetime
    :param str_time: Specify the datetime in a string format.
    :return: Date and time in  a datetime format.
    """
    UTC_TIME = strp_dtime(str_time).replace(tzinfo=timezone("UTC"))
    return UTC_TIME.astimezone(tz=timezone("Asia/Calcutta")).replace(tzinfo=None)


def check_file(file, _json=True):
    """
    This method will check if the file exist. If not it will create it.
    :param file: Specify the file name
    :param _json: JSON Bool will decide if file is of type JSON or plain text file.
    :return: None
    """
    try:
        read(file)
    except FileNotFoundError:
        if _json:
            write(file, {})
        else:
            write(file, "")


def get_unique_id():
    """
    This method will give us a unique id.
    :return:
    """
    return str(uuid.uuid4().fields[-1])


def cinput(string):
    """
    This method will validate the inputs. If input is blank it will not accept it.
    :param string: Specify input string.
    :return:
    """
    is_field_blank = True
    while is_field_blank:
        field = input(string)
        if field == "":
            print("field cannot be blank.")
        else:
            is_field_blank = False
            return field


def get_req(endpoint, org, req_id):
    """
    This method acts as base request method and will perform GET requests for a given endpoint.
    :param req_id: Specify Unique ID to distinguish the requests
    :param org: Specify the organisation
    :param endpoint: Specify the endpoint
    :return: Result of the request
    """

    # Set a flag to not complete
    global request, request_completed, response, response_error
    request_completed = False
    while not request_completed:
        try:
            headers = org["headers"]
            url = f"{org['url']}/{endpoint}"
            logger.info(f"URL: {url}")

            # if response is received than set flag to complete
            request = requests.get(url=url, headers=headers, timeout=15)
            time.sleep(1)
            response = request.json()
            response_error = request.text
            if validate_request(response, req_id):
                request_completed = True
                logger.info(
                    f"Req ID: {req_id} Request Completed --> endpoint: /{endpoint}, organisation: {org['org_id']} "
                    f"({org['name']})")
                return response
            else:
                logger.warning(
                    f"Req ID: {req_id} Request Incomplete --> endpoint: /{endpoint}, organisation: {org['org_id']} "
                    f"({org['name']})")
                time.sleep(20)
                request_completed = False

        # if error occurs than try again until its complete.
        except (ConnectionError, requests.exceptions.ConnectionError,
                requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ReadTimeout, ConnectionResetError):
            logger.exception(
                f"Req ID: {req_id} Request Incomplete --> endpoint: /{endpoint}, organisation: {org['org_id']} "
                f"({org['name']}) "
                f"Error: ConnectionError Occurred:\n"
                f" Suspected reasons--> No Internet Connectivity, Dashboard API not responding.")
            request_completed = False
            time.sleep(20)

        except (json.JSONDecodeError,):
            logger.exception(f"Req ID: {req_id} Request Incomplete (JSON NO Response or INVALID Response) --> "
                             f"endpoint: /{endpoint}, "
                             f"organisation: {org['org_id']} ({org['name']})")
            request_completed = False
            time.sleep(20)

        except (TypeError,):
            logger.exception(f"Req ID: {req_id} Request Incomplete (Type Error Response) --> endpoint: /{endpoint}, "
                             f"organisation: {org['org_id']} ({org['name']})")
            request_completed = False
            time.sleep(20)


def validate_request(req, req_id):
    """
    This method will check if request pass the defined validations.
    :param req_id: unique id
    :param req: Specify request
    :return: Bool
    """
    if req is None:
        logger.error(f"Req ID: {req_id} Request failed validation, Req was None, Req: {req}")
        return False
    elif req == {}:
        logger.error(f"Req ID: {req_id} Request failed validation, Req was empty json, Req: {req}")
        return False
    elif req == "":
        logger.error(f"Req ID: {req_id} Request failed validation, Req was blank, Req: {req}")
        return False
    elif req == {"errors": ["API rate limit exceeded for organization"]}:
        logger.error(f"Req ID: {req_id} Request failed validation, API rate limit exceeded, Req: {req}")
        logger.info("Max API Limit reached")
        return False
    elif req == {'errors': ['Invalid API key']}:
        logger.error(f"Req ID: {req_id} Request failed validation, Invalid API Key, Req: {req}")
        logger.info("Invalid Api Key")
        return False
    else:

        return True


def get_org_devices_status(org):
    """
    This method will get the status of all the devices in an organization.
    :return: list of device info and status.
    """
    uid = get_unique_id()
    org["url"] = org["url"].replace("v1", "v0")
    logger.info(
        f"Req ID: {uid} Request: get_org_devices_status, endpoint: /deviceStatuses, organisation: {org['org_id']} "
        f"({org['name']})")
    return get_req("deviceStatuses", org, uid)


def get_networks(org):
    """
    This method will get all networks in an organization.
    :return: list of networks
    """
    uid = get_unique_id()
    org["url"] = org["url"].replace("v1", "v0")
    logger.info(f"Req ID: {uid} Request: get_networks, endpoint: /networks, organisation: {org['org_id']} "
                f"({org['name']})")
    return get_req("networks", org, uid)


def get_network_devices(networkId, org):
    """
    This method will get all devices present in the network.
    :param org: Specify Organisation
    :param networkId: Specify the network id.
    :return: list of devices
    """
    uid = get_unique_id()
    logger.info(
        f"Req ID: {uid} Request: get_network_devices, endpoint: /networks/{networkId}/devices, organisation: {org['org_id']} "
        f"({org['name']})")
    return get_req(f"networks/{networkId}/devices", org, uid)


def get_all_devices_dep(org):
    """
    This method will get all devices in organization
    :return: list of devices
    """
    uid = get_unique_id()
    logger.info(f"{uid} Request: get_all_devices ")
    return get_req("devices?total_pages=all", org, uid)


def get_org_devices_status_of_both_dashboard_dep():
    # Deprecated..........................
    logger.info("Request: get_org_devices_status_of_both_dashboard_dep ")
    # possibilities = [True, False]
    possibilities = [False]
    global_devices_status_list = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        thread = [executor.submit(get_org_devices_status)]
        # thread = [executor.submit(get_network_devices, network["id"]) for network in list(read(ORG_NETWORKS_FILE))]
        for output in concurrent.futures.as_completed(thread):
            org_status: list = output.result()
            for device in org_status:
                global_devices_status_list.append(device)

    return global_devices_status_list


def get_org_devices_status_of_all_dashboard():
    """
    This method will return organisation status of all organisations
    :return: list
    """
    all_org_device_list = []
    for org in list(read(CONFIG_FILE)):
        org_device_status = get_org_devices_status(org)
        time.sleep(1)
        for device in org_device_status:
            all_org_device_list.append(device)
    """with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        thread = [executor.submit(get_org_devices_status, org) for org in list(read(CONFIG_FILE))]
        # thread = [executor.submit(get_network_devices, network["id"]) for network in list(read(ORG_NETWORKS_FILE))]
        for output in concurrent.futures.as_completed(thread):
            org_status: list = output.result()
            for device in org_status:
                all_org_device_list.append(device)"""

    return all_org_device_list


def get_org_devices_status_of_all_dashboard_for_DB():
    """
    This method will return organisation status of all organisations
    :return: list
    """
    logger.info("Request: get_org_devices_status_of_all_dashboard ")
    DEVICES_STATUS_COLLECTION.create_index([("serial", ASCENDING)], unique=True)
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        thread = [executor.submit(get_org_devices_status, org) for org in ORG_CONFIG_COLLECTION.find()]
        # thread = [executor.submit(get_network_devices, network["id"]) for network in list(read(ORG_NETWORKS_FILE))]
        for output in concurrent.futures.as_completed(thread):
            org_status: list = output.result()
            for device in org_status:
                try:
                    DEVICES_STATUS_COLLECTION.insert_one(device)
                except DuplicateKeyError:
                    device_db = DEVICES_STATUS_COLLECTION.find_one(
                        {"serial": device["serial"]})
                    remove_id = device.pop("_id")
                    DEVICES_STATUS_COLLECTION.find_one_and_update({"_id": ObjectId(device_db["_id"])}, {"$set": device})
                    # DEVICES_STATUS_COLLECTION.find_one_and_update({"_id": device_db["_id"]["$oid"]}, {"$set": device})
    logger.info("Devices status of all Org for DB completed")
    return DEVICES_STATUS_COLLECTION.find()


def create_active_network_devices_file():
    """
    This method will create a dictionary of devices.
    :return: dictionary of devices.
    """
    logger.info("Gathering Active network devices..")
    all_networks = []
    # create an empty list
    global_devices_list = []
    # save networks info and devices info in a file.
    for org in read(CONFIG_FILE):
        org: dict
        networks: list = get_networks(org)
        if org["china_dashboard"] is True:
            networks[0]["tags"] = str(networks[0]["tags"]) + " 0-Active"

        """for network in networks:
            network["tags"].append("0-Active")"""
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            """
            # Global use
            thread = [executor.submit(get_network_devices, network["id"], org) for network in networks
                      if "0-Active" in network["tags"]]
            """

            # Specific to Project Use.
            thread = [executor.submit(get_network_devices, network["id"], org) for network in networks
                      if "0-Active" in network["tags"]]
            for output in concurrent.futures.as_completed(thread):
                for device in output.result():
                    global_devices_list.append(device)
        for network in networks:
            all_networks.append(network)
    write(ORG_DEVICE_STATUS_FILE, get_org_devices_status_of_all_dashboard())
    write(ORG_DEVICES, global_devices_list)
    write(ORG_NETWORKS_FILE, all_networks)


def create_active_network_devices_file_for_DB():
    """
    This method will create a dictionary of devices.
    :return: dictionary of devices.
    """
    logger.info("Gathering Active network devices..")
    all_networks = []
    # create an empty list
    global_devices_list = []
    # save networks info and devices info in a file.
    for org in ORG_CONFIG_COLLECTION.find():
        networks = get_networks(org)

        if org["china_dashboard"] is True:
            networks[0]["tags"].append("0-Active")

        for network in networks:
            all_networks.append(network)

        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            thread = [executor.submit(get_network_devices, network["id"], org) for network in networks]
            for devices in concurrent.futures.as_completed(thread):
                time.sleep(0.1)
                for device in devices.result():
                    global_devices_list.append(device)

    get_org_devices_status_of_all_dashboard_for_DB()

    # Add data to MongoDB
    NETWORKS_COLLECTION.create_index([("networkId", ASCENDING)], unique=True)
    NETWORK_MAINTENANCE_COLLECTION.create_index([("networkId", ASCENDING)], unique=True)
    DEVICES_STATUS_COLLECTION.create_index([("serial", ASCENDING)], unique=True)
    DEVICES_COLLECTION.create_index([("serial", ASCENDING)], unique=True)

    for network in all_networks:
        network["networkId"] = network.pop("id")
        # -------------------------------#
        # Add Networks in ORG_NETWORK_DB #
        # -------------------------------#

        try:
            NETWORKS_COLLECTION.insert_one(network)
        except DuplicateKeyError:
            network_db = NETWORKS_COLLECTION.find_one(
                {"networkId": network["networkId"]})
            remove_id = network.pop("_id")
            NETWORKS_COLLECTION.find_one_and_update({"_id": ObjectId(network_db["_id"])}, {"$set": network})

        # -------------------------------------------#
        # Add Networks in ORG_NETWORK_MAINTENANCE_DB #
        # -------------------------------------------#
        try:
            NETWORK_MAINTENANCE_COLLECTION.insert_one(network)
        except DuplicateKeyError:
            network_db = NETWORK_MAINTENANCE_COLLECTION.find_one(
                {"networkId": network["networkId"]})
            remove_id = network.pop("_id")
            NETWORK_MAINTENANCE_COLLECTION.find_one_and_update({"_id": ObjectId(network_db["_id"])}, {"$set": network})

    # -------------------------------#
    # Add Networks in ORG_DEVICES_DB #
    # -------------------------------#
    for device in global_devices_list:
        try:
            DEVICES_COLLECTION.insert_one(device)
        except DuplicateKeyError:
            device_db = DEVICES_COLLECTION.find_one(
                {"serial": device["serial"]})
            remove_id = device.pop("_id")
            DEVICES_COLLECTION.find_one_and_update({"_id": ObjectId(device_db["_id"])}, {"$set": device})

    logger.info("Active network devices file creation for DB completed")


def month_folder_name():
    """
    This method will return a name of the folder in "PresentMonth PresentYear" format eg. "March 2021".
    :return: name of the folder.
    """
    get_datetime = datetime.now()
    format_time = get_datetime.strftime("%B %Y")
    return format_time


def create_folder_every_month():
    """
    This method will create a folder
    :return: None
    """

    EVERY_MONTH_REPORT_PATH = os.path.join(MONTHLY_REPORT_PATH, month_folder_name())
    if not os.path.exists(EVERY_MONTH_REPORT_PATH):
        os.mkdir(EVERY_MONTH_REPORT_PATH)
        with open(f"{EVERY_MONTH_REPORT_PATH}/report_not_yet_generated", "w") as f:
            pass


def check_for_twelve(p_time):
    """
    This method will check if the time is between the specified range.
    :param p_time: Specify the current time.
    :return: boolean
    """

    if datetime.strptime("23:55:00", "%H:%M:%S").time() < p_time < datetime.strptime("23:59:00", "%H:%M:%S").time():
        logger.info("Time 23:55 has been passed. Time to copy monthly report...")
        return True
    else:
        return False


def check_for_clock_daily(p_time):
    """
    This method will check if the time is between the specified range.
    :param p_time: Specify the current time.
    :return: boolean
    """

    if datetime.strptime("22:15:00", "%H:%M:%S").time() < p_time < datetime.strptime("23:59:00", "%H:%M:%S").time():
        logger.info("Time 22:15 has been passed. Time to copy the daily report..")
        return True
    else:
        return False


def check_leap_year(year):
    """
    This method will check if specified year is a leap year or not.
    :param year: Specify the year
    :return: boolean
    """

    # check if year is divisible by 4
    if (year % 4) == 0:
        # if yes check if it is divisible by 100
        if (year % 100) == 0:
            # if yes check if it is divisible by 400
            if (year % 400) == 0:
                # if all conditions passed than it is a leap year.
                return True
            else:
                return False
        else:
            return False
    else:
        return False


def last_day(p_datetime: datetime):
    """
    This method will return the last day of the month.
    :param p_datetime: Specify the date and time.
    :return: Last day
    """
    month = p_datetime.month
    year = p_datetime.year
    if month == 2:
        if check_leap_year(year) is True:
            return 29
        else:
            return 28
    else:
        for mapping in LAST_DAYS_MAPPING:
            map_month, map_day = mapping[0], mapping[1]
            if month == map_month:
                return map_day


def copy_report():
    """
    This method will copy the report on last day of every month between 23:55 - 23:59.
    :return: None
    """

    EVERY_MONTH_REPORT_PATH = os.path.join(MONTHLY_REPORT_PATH, month_folder_name())
    if not os.path.exists(f"{EVERY_MONTH_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}"):
        present_datetime = datetime.now()
        if present_datetime.day == last_day(present_datetime) and check_for_twelve(datetime.now().time()) is True:
            try:
                shutil.copy2(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}",
                             f"{MONTHLY_REPORT_PATH}/{month_folder_name()}/{SINGLE_AVAILABILITY_REPORT_NAME}")

                send_notification_email(subject=f"[v0.2] Meraki Availability Report for {month_folder_name()}",
                                        html_content=SEND_REPORT_CONTENT,
                                        attachment=LIVE_REPORT_FILE, filename="Availability_Report.xlsx",
                                        password=get_password())
                try:
                    os.remove(f"{EVERY_MONTH_REPORT_PATH}/report_not_yet_generated")
                except FileNotFoundError:
                    logger.error("No file named 'report_not_yet_generated' in the directory")
                logger.info("Removing Files Directory for fresh restart")
                shutil.rmtree(FILES_PATH)
                logger.info("*********************************************************")
                logger.info(f"Monthly Report Copied to {month_folder_name()} folder")
                logger.info("*********************************************************")
                logger.info("Sleeping for 10 minutes")
                time.sleep(600)
            except PermissionError:
                logger.error("PermissionError, Report is open. It will now be updated in the next execution")
    return


def copy_report_for_DB():
    """
    This method will copy the report on last day of every month between 23:55 - 23:59.
    :return: None
    """

    EVERY_MONTH_REPORT_PATH = os.path.join(MONTHLY_REPORT_PATH, month_folder_name())
    if not os.path.exists(f"{EVERY_MONTH_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}"):
        present_datetime = datetime.now()
        if present_datetime.day == last_day(present_datetime) and check_for_twelve(datetime.now().time()) is True:
            try:
                shutil.copy2(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}",
                             f"{MONTHLY_REPORT_PATH}/{month_folder_name()}/{SINGLE_AVAILABILITY_REPORT_NAME}")
                try:
                    os.remove(f"{EVERY_MONTH_REPORT_PATH}/report_not_yet_generated")
                except FileNotFoundError:
                    logger.error("No file named 'report_not_yet_generated' in the directory")
                logger.info("Creating backup of sla's for all devices")
                create_backup_of_sla_collection()
                logger.info("Deleting the SLA Table for resetting the start time of all devices")
                SLA_COLLECTION.drop()
                logger.info("*********************************************************")
                logger.info(f"Monthly Report Copied to {month_folder_name()} folder")
                logger.info("*********************************************************")
                logger.info("Sleeping for 10 minutes")
                time.sleep(600)
            except PermissionError:
                logger.error("PermissionError, Report is open. It will now be updated in the next execution")
    return


def copy_report_every_day():
    """
    This method will copy the report everyday between 23:50 - 23:59.
    :return: None
    """
    present_datetime = datetime.now()
    DAILY_REPORT_NAME = f"Availability Report {present_datetime.strftime('%d_%m_%Y')}.xlsx"
    if not os.path.exists(f"{DAILY_REPORT_PATH}/{DAILY_REPORT_NAME}"):

        if check_for_clock_daily(datetime.now().time()) is True:
            try:
                shutil.copy2(f"{LIVE_REPORT_PATH}/{SINGLE_AVAILABILITY_REPORT_NAME}",
                             f"{DAILY_REPORT_PATH}/{DAILY_REPORT_NAME}")
                logger.info(f"Daily Report Copied to Daily Reports folder")
            except PermissionError:
                logger.error("PermissionError, Report is open. It will now be updated in the next execution")
    return


def create_header_cell(row, column, value, target_sheet):
    """
    This method will create a header cell in excel sheet
    :param row: Specify the row number
    :param column: Specify the column number
    :param value: Specify the value
    :param target_sheet: Specify the excel sheet
    :return: None
    """
    font = Font(bold=True, size=15)
    fill = PatternFill(fill_type="solid", fgColor="4E94CF")
    thick = Side(border_style="thick", color='000000')
    border = Border(left=thick, right=thick, top=thick, bottom=thick)
    cell = target_sheet.cell(row=row, column=column)
    cell.value = value
    cell.fill = fill
    cell.border = border
    cell.font = font


def create_data_cell(row, column, value, target_sheet):
    """
    This method will create a data cell in excel sheet
    :param row: Specify the row number
    :param column: Specify the column number
    :param value: Specify the value
    :param target_sheet: Specify the excel sheet
    :return: None
    """
    font = Font(bold=False, size=12)
    fill = PatternFill(fill_type="solid", fgColor="F5F5F5")
    thick = Side(border_style="thin", color='000000')
    border = Border(left=thick, right=thick, top=thick, bottom=thick)
    cell = target_sheet.cell(row=row, column=column)
    cell.value = value
    cell.fill = fill
    cell.border = border
    cell.font = font


# Initial configurations to create config file
def get_meraki_credentials():
    """
    This method will ask for organisation id, api key and url.
    :return:
    """
    config = {}
    org_id = cinput("Enter Organisation ID: ")
    name = cinput("Enter the Organisation name (Just for recognition purpose. You can set any name): ")
    api_key = cinput("Enter API Key: ")
    base_url = input("Enter Base URL [default: https://api.meraki.com/api/v1/organizations]: ")
    tags = input("Want to add specific tags that you want to scan [default: None]: ")
    if base_url == "":
        base_url = "https://api.meraki.com/api/v1/organizations"
    base_url = f"{base_url}/{org_id}"

    if tags == "":
        tags = []
    else:
        tags.split(",")
    if "meraki.cn" in base_url:
        china_dashboard = True
    else:
        china_dashboard = False
    headers = {"Content-Type": "application/json", "Accept": "application/json", "X-Cisco-Meraki-API-Key": api_key}
    return {"org_id": org_id, "name": name, "api_key": api_key, "url": base_url, "china_dashboard": china_dashboard,
            "tags": tags, "headers": headers}


def make_config():
    """
    This method will make the config file to start the script
    :return:
    """
    organisations = []
    flag = True
    while flag:
        organisations.append(get_meraki_credentials())
        x = cinput("Organisation Added. Do you want to add more [y] or [n]: ")
        if x == "n":
            flag = False
    write(CONFIG_FILE, organisations)
    setup_email()


def setup_email():
    generate_key()
    # Setup Email account.
    print("Setup up Email account for sending notifications")
    email_address = input(f"Enter the email address [default: {EMAIL_ADDRESS}]: ")
    if email_address == "":
        email_address = EMAIL_ADDRESS
    password = cinput(f"Password: ")
    credentials = {"email": email_address, "password": encrypt_message(password).decode("ascii")}
    write(EMAIL_CREDENTIALS, credentials)


def get_password():
    EMAIL_PASSWORD = decrypt_message(read(EMAIL_CREDENTIALS)["password"].encode())
    return EMAIL_PASSWORD


def make_config_for_DB():
    """
    This method will make the config file to start the script
    :return:
    """
    ORG_CONFIG_COLLECTION.create_index([("org_id", ASCENDING)], unique=True)
    organisations = []
    flag = True
    while flag:
        organisations.append(get_meraki_credentials())
        x = input("Organisation Added. Do you want to add more [y] or [n]: ")
        if x == "n":
            flag = False

    for org in organisations:
        try:
            ORG_CONFIG_COLLECTION.insert_one(org)
        except DuplicateKeyError:
            db_org = ORG_CONFIG_COLLECTION.find_one({"org_id": org["org_id"]})
            org.pop("_id")
            ORG_CONFIG_COLLECTION.find_one_and_update({"org_id": db_org["org_id"]}, {"$set": org})

    return ORG_CONFIG_COLLECTION.find()


def ask_initial_config_for_DB():
    """
    This method will check for config file. If not exist it will create it
    :return: None
    """
    try:
        if ORG_CONFIG_COLLECTION.estimated_document_count() != 0:
            logger.info("Already have config file")
            time.sleep(0.1)
            logger.info(f"Parameters are as follows:")
            time.sleep(0.1)
            for org in ORG_CONFIG_COLLECTION.find():
                org: dict = org
                logger.info(f"| Org ID: {org['org_id']} "
                            f"| Api Key: {org['api_key']} "
                            f"| Url: {org['url']}")
                time.sleep(0.1)
            ask = input("Do you want to edit the config file [y] or do you want to continue with same configs [n]: ")
            if ask == "y":
                make_config_for_DB()
            elif ask == "n":
                pass
        elif ORG_CONFIG_COLLECTION.estimated_document_count() == 0:
            logger.info("No config file found. Please fill below details to create one.")
            make_config_for_DB()

    except FileNotFoundError:
        logger.info("No config file found. Please fill below details to create one.")
        make_config_for_DB()


def ask_initial_config():
    """
    This method will check for config file. If not exist it will create it
    :return: None
    """
    try:
        config_file = read(CONFIG_FILE)
        logger.info("Already have config file")
        time.sleep(0.1)
        logger.info(f"Parameters are as follows:")
        time.sleep(0.1)
        for org in config_file:
            org: dict = org
            logger.info(f"| Org ID: {org['org_id']} "
                        f"| Api Key: {org['api_key']} "
                        f"| Url: {org['url']}")
            time.sleep(0.1)
        ask = cinput("Do you want to edit the config file [y] or do you want to continue with same configs [n]: ")
        if ask == "y":
            make_config()
        elif ask == "n":
            pass

    except FileNotFoundError:
        logger.info("No config file found. Please fill below details to create one.")
        make_config()


def create_thread(function, items, item=None):
    # Not used anywhere yet.
    with concurrent.futures.ThreadPoolExecutor() as executor:
        if item is not None:
            thread = [executor.submit(function, item) for item in items]
        else:
            thread = [executor.submit(function) for _ in items]
        for output in concurrent.futures.as_completed(thread):
            pass


def device_in_mapper_already_exists(serial):
    """
    This method will check if the device exist in Mapper Collection. (Used in Database operation)
    :param serial: Specify device serial number
    :return: Bool
    """
    if MAPPER_COLLECTION.find_one({"serial": serial}) is not None:
        return True
    else:
        return False


def create_backup_of_sla_collection():
    """
    This method will create a backup of 'org_device_status' collection every month
    :return: String
    """
    backup = []
    unique_string = f"Backup for Month {month_folder_name()}"
    backup_template = {}
    BACKUP_SLA_COLLECTION.create_index([("backup_info", ASCENDING)], unique=True)
    for document in SLA_COLLECTION.find():
        backup.append(document)
    backup_template.update({"name": unique_string,
                            "backup": backup})
    try:
        BACKUP_SLA_COLLECTION.insert_one(backup_template)
    except DuplicateKeyError:
        pass
    return f"Backup Created for {month_folder_name()}"


def get_device(serial):
    for device in read(ORG_DEVICES):
        device: dict
        if device["serial"] == serial:
            return device


def add_tag_manually_to_network(networkId, tag):
    networks = read(ORG_NETWORKS_FILE)
    for network in networks:
        network: dict
        if network["id"] == networkId:
            network["tags"].append(tag)
    write(ORG_NETWORKS_FILE, networks)
    for network in networks:
        network: dict
        if network["id"] == networkId:
            return network


def remove_tag_manually_from_network(networkId, tag):
    networks = read(ORG_NETWORKS_FILE)
    for network in networks:
        network: dict
        if network["id"] == networkId:
            network["tags"].remove(tag)
    write(ORG_NETWORKS_FILE, networks)


def generate_key():
    """
    Generates a key and save it into a file
    """
    key = Fernet.generate_key()
    with open(SECRET_KEY_FILE, "wb") as key_file:
        key_file.write(key)


def load_key():
    """
    Load the previously generated key
    """
    return open(SECRET_KEY_FILE, "rb").read()


def encrypt_message(message):
    """
    Encrypts a message
    """
    key = load_key()
    encoded_message = message.encode()
    f = Fernet(key)
    encrypted_message = f.encrypt(encoded_message)

    return encrypted_message


def decrypt_message(encrypted_message):
    """
    Decrypts an encrypted message
    """
    key = load_key()
    f = Fernet(key)
    decrypted_message = f.decrypt(encrypted_message)

    return decrypted_message.decode().replace("'", '"')
